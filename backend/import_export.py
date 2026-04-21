"""
Import CSV / Export Excel + PDF
"""

import io
import csv
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from database.models import Job, Machine, Operation, Planning, PlanningOperation


# ══════════════════════════════════════════════════════════════════════════════
# IMPORT CSV
# ══════════════════════════════════════════════════════════════════════════════

def import_jobs_csv(db, content: str) -> dict:
    """
    Importe des jobs depuis un CSV.
    Format attendu : nom,priorite,date_due,statut
    """
    reader = csv.DictReader(io.StringIO(content))
    imported = 0
    errors = []

    for i, row in enumerate(reader, start=2):
        try:
            nom = row.get("nom", "").strip()
            if not nom:
                errors.append(f"Ligne {i} : nom manquant")
                continue

            priorite = int(row.get("priorite", 2))
            statut = row.get("statut", "en_attente").strip()
            date_due_str = row.get("date_due", "").strip()
            date_due = None
            if date_due_str:
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M"):
                    try:
                        date_due = datetime.strptime(date_due_str, fmt)
                        break
                    except ValueError:
                        continue

            job = Job(nom=nom, priorite=priorite, statut=statut, date_due=date_due)
            db.session.add(job)
            imported += 1

        except Exception as e:
            errors.append(f"Ligne {i} : {str(e)}")

    db.session.commit()
    return {"imported": imported, "errors": errors}


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════

BLUE = "2563EB"
LIGHT_BLUE = "EFF6FF"
HEADER_GRAY = "F1F5F9"

def _header_style(ws, row, headers):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def export_excel(db) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # ── Feuille Jobs ──────────────────────────────────────────────────────────
    ws_jobs = wb.active
    ws_jobs.title = "Jobs"
    _header_style(ws_jobs, 1, ["ID", "Nom", "Priorité", "Date Échéance", "Statut", "Nb Opérations", "Durée Totale (min)"])
    for row_idx, job in enumerate(Job.query.all(), start=2):
        ws_jobs.append([
            job.id, job.nom, job.priorite,
            job.date_due.strftime("%d/%m/%Y") if job.date_due else "—",
            job.statut,
            len(job.operations),
            sum(o.duree for o in job.operations),
        ])
        if row_idx % 2 == 0:
            for col in range(1, 8):
                ws_jobs.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    # ── Feuille Machines ──────────────────────────────────────────────────────
    ws_m = wb.create_sheet("Machines")
    _header_style(ws_m, 1, ["ID", "Nom", "Type", "Capacité", "Statut", "Temps Setup (min)"])
    for m in Machine.query.all():
        ws_m.append([m.id, m.nom, m.type, m.capacite, m.statut, m.temps_setup])

    # ── Feuille Opérations ────────────────────────────────────────────────────
    ws_ops = wb.create_sheet("Opérations")
    _header_style(ws_ops, 1, ["ID", "Job", "Machine", "Durée (min)", "Ordre", "Statut"])
    for op in Operation.query.all():
        ws_ops.append([
            op.id,
            op.job.nom if op.job else "—",
            op.machine.nom if op.machine else "—",
            op.duree, op.ordre, op.statut,
        ])

    # ── Feuille Plannings ─────────────────────────────────────────────────────
    ws_pl = wb.create_sheet("Plannings")
    _header_style(ws_pl, 1, ["ID", "Nom", "Algorithme", "Makespan (min)", "Taux Retard (%)", "Date Création"])
    for pl in Planning.query.order_by(Planning.date_creation.desc()).all():
        ws_pl.append([
            pl.id, pl.nom, pl.algorithme, pl.makespan,
            round(pl.taux_retard * 100, 1),
            pl.date_creation.strftime("%d/%m/%Y %H:%M"),
        ])

    # Ajuster largeurs
    for ws in [ws_jobs, ws_m, ws_ops, ws_pl]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT PDF
# ══════════════════════════════════════════════════════════════════════════════

def export_pdf_report(db) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  fontSize=20, textColor=colors.HexColor("#2563EB"),
                                  spaceAfter=12, alignment=TA_CENTER)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
                               fontSize=13, textColor=colors.HexColor("#0F172A"),
                               spaceBefore=16, spaceAfter=6)

    story = []
    story.append(Paragraph("📊 Rapport de Production — Job Shop Scheduler", title_style))
    story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    def make_table(data, col_widths=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8F9FC"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    # ── Jobs ──────────────────────────────────────────────────────────────────
    story.append(Paragraph("Jobs de Production", h2_style))
    jobs = Job.query.all()
    job_data = [["ID", "Nom", "Priorité", "Date Échéance", "Statut", "Durée (min)"]]
    for j in jobs:
        job_data.append([
            j.id, j.nom[:35], j.priorite,
            j.date_due.strftime("%d/%m/%Y") if j.date_due else "—",
            j.statut,
            sum(o.duree for o in j.operations),
        ])
    story.append(make_table(job_data, [1.5*cm, 8*cm, 2*cm, 3.5*cm, 3*cm, 3*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── Machines ──────────────────────────────────────────────────────────────
    story.append(Paragraph("Machines", h2_style))
    machines = Machine.query.all()
    m_data = [["ID", "Nom", "Type", "Capacité", "Statut", "Setup (min)"]]
    for m in machines:
        m_data.append([m.id, m.nom, m.type, m.capacite, m.statut, m.temps_setup])
    story.append(make_table(m_data, [1.5*cm, 7*cm, 4*cm, 2.5*cm, 4*cm, 2.5*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── Plannings ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Historique Plannings", h2_style))
    plannings = Planning.query.order_by(Planning.date_creation.desc()).limit(10).all()
    pl_data = [["ID", "Nom", "Algorithme", "Makespan (min)", "Taux Retard (%)", "Date"]]
    for pl in plannings:
        pl_data.append([
            pl.id, pl.nom[:40], pl.algorithme,
            round(pl.makespan, 1),
            f"{pl.taux_retard * 100:.1f}%",
            pl.date_creation.strftime("%d/%m/%Y %H:%M"),
        ])
    if len(pl_data) > 1:
        story.append(make_table(pl_data, [1.5*cm, 9*cm, 3*cm, 3*cm, 3*cm, 3.5*cm]))

    doc.build(story)
    buf.seek(0)
    return buf
